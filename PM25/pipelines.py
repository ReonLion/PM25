# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html
import time
from openpyxl import Workbook, load_workbook
import openpyxl
import os
from PM25.items import *
import matplotlib.pyplot as plt
from matplotlib.font_manager import *

class Pm25Pipeline(object):
    def process_item(self, item, spider):
        file_dir = os.path.join(os.getcwd(), item['check_time'], '')
        
        if os.path.isdir(file_dir):
            pass
        else:
            os.mkdir(file_dir)
            
        if spider.name == 'CitiesRank':
            if isinstance(item, CitiesRankItem):
                excel_name = os.path.join(file_dir, u'实时排行.xlsx')
                self.save_to_excel(item, excel_name)
            elif isinstance(item, CityLinechartItem):
                self.draw_line_chart(item, spider, file_dir)
            elif isinstance(item, RelevantCityItem):
                self.save_to_city_excel(item, spider, file_dir)
            elif isinstance(item, CityItem):
                excel_path = os.path.join(file_dir, item['city'], '')
                if not os.path.exists(excel_path):
                    os.mkdir(excel_path)
                excel_name = os.path.join(excel_path, item['city']+'.xlsx')
                self.save_to_excel(item, excel_name)
                
        elif spider.name == 'CitiesRank1day':
            excel_name = os.path.join(file_dir, u'昨日排行.xlsx')
            self.save_to_excel(item, excel_name)
        elif spider.name == 'CitiesRank7day':
            excel_name = os.path.join(file_dir, u'近七日排行.xlsx')
            self.save_to_excel(item, excel_name)
        elif spider.name == 'CitiesRank30day':
            excel_name = os.path.join(file_dir, u'上月排行.xlsx')
            self.save_to_excel(item, excel_name)
        
        return item
    
    def save_to_excel(self, item, excel_name):
        if not os.path.exists(excel_name):
            wb = Workbook()
            wb.save(excel_name)
        wb = load_workbook(excel_name)
        
        try:
            ws = wb.get_sheet_by_name('sheet1')
        except:
            ws = wb.create_sheet(title='sheet1', index=0)
            '''
            65为A的ASCII码
            将item的各项key存入A1,B1,C1等
            '''
            i = 65
            for key in item:
                ws[chr(i)+'1'] = key
                i += 1
        
        last_row = ws.max_row+1
        for i in range(1, len(item)+1):
            ws.cell(row=last_row, column=i).value = item[ws.cell(row=1, column=i).value]
        wb.save(excel_name)
        
    def draw_line_chart(self, item, spider, file_dir):
        # 显示中文正常
        myfont = FontProperties(fname='/usr/share/fonts/truetype/arphic/ukai.ttc')
        # 显示负号正常
        matplotlib.rcParams['axes.unicode_minus'] = False
        
        # 设置ｘ轴的实际坐标为1,2,3...30
        x_30d_list = list(range(30))
        x_24h_list = list(range(24))
        matplotlib.style.use('ggplot')
        # 设置画布大小
        plt.figure(figsize=(10, 5))
        # 画两条线
        plt.plot(x_30d_list, item['y_aqi_china_30d'], label='china')
        plt.plot(x_30d_list, item['y_aqi_usa_30d'], label='usa')
        # ｘ设计坐标与显示坐标...日的映射
        plt.xticks(x_30d_list, self.list_encode(item['x_time_30d']),fontproperties=myfont)
        # 设置x轴与y轴的标签
        plt.xlabel(u'30天', fontproperties=myfont)
        plt.ylabel(u'aqi', fontproperties=myfont)
        # 设置y轴从零开始
        plt.ylim(ymin=0)
        # 设置label显示生效
        plt.legend()
        # 保持图像
        png_path = os.path.join(file_dir, item['city'], '')
        if not os.path.exists(png_path):
            os.mkdir(png_path)
        png_name_30d = os.path.join(png_path,'30d.png')
        plt.savefig(png_name_30d, bbox_inches='tight')
        plt.close()
        
        plt.figure(figsize=(10, 5))
        plt.plot(x_24h_list, item['y_aqi_china_24h'], label='china')
        plt.plot(x_24h_list, item['y_aqi_usa_24h'], label='usa')
        plt.xticks(x_24h_list, self.list_encode(item['x_time_24h']),fontproperties=myfont)
        plt.xlabel(u'24小时', fontproperties=myfont)
        plt.ylabel(u'aqi', fontproperties=myfont)
        plt.ylim(ymin=0)
        plt.legend()
        png_path = os.path.join(file_dir, item['city'], '')
        if not os.path.exists(png_path):
            os.mkdir(png_path)
        png_name_24h = os.path.join(png_path,'24h.png')
        plt.savefig(png_name_24h, bbox_inches='tight')
        plt.close()
        
        excel_path = os.path.join(file_dir, item['city'], '')
        if not os.path.exists(excel_path):
            os.mkdir(excel_path)
        excel_name = os.path.join(excel_path, item['city']+'.xlsx')
        if not os.path.exists(excel_name):
            wb = Workbook()
            wb.save(excel_name)
        wb = load_workbook(excel_name)
        
        try:
            ws = wb.get_sheet_by_name('24h')
        except:
            ws = wb.create_sheet(title='24h', index=2)
        img = openpyxl.drawing.image.Image(png_name_24h)
        ws.add_image(img, 'A1')
        
        try:
            ws = wb.get_sheet_by_name('30d')
        except:
            ws = wb.create_sheet(title='30d', index=3)
        # 将图像插入到表格
        img = openpyxl.drawing.image.Image(png_name_30d)
        ws.add_image(img,'A1')
        wb.save(excel_name)
        
    def list_encode(self, list_example):
        for i in list_example:
            yield i.decode('utf-8')
            
    def save_to_city_excel(self, item, spider, file_dir):
        excel_path = os.path.join(file_dir, item['city'], '')
        if not os.path.exists(excel_path):
            os.mkdir(excel_path)
        excel_name = os.path.join(excel_path, item['city']+'.xlsx')
        if not os.path.exists(excel_name):
            wb = Workbook()
            wb.save(excel_name)
        wb = load_workbook(excel_name)
        try:
            ws = wb.get_sheet_by_name(u'相关城市')
        except:
            ws = wb.create_sheet(title=u'相关城市', index=1)
            '''
            65为A的ASCII码
            将item的各项key存入A1,B1,C1等
            '''
            i = 65
            for key in item:
                ws[chr(i)+'1'] = key
                i += 1
                
        last_row = ws.max_row+1
        for i in range(1, len(item)+1):
            ws.cell(row=last_row, column=i).value = item[ws.cell(row=1, column=i).value]
        wb.save(excel_name)